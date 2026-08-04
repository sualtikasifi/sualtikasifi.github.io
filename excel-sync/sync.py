#!/usr/bin/env python3
"""
Marder Ciftlik - Buzagi Tedavileri Excel <-> Site senkronizasyonu.

Iki yonlu calisir:
  - Excel'de elle girilen yeni tedaviler siteye (Supabase) aktarilir.
  - Sitede eklenen yeni tedaviler Excel'e satir olarak eklenir.
  - Ayni kaydin hem Excel'de hem sitede DEGISTIRILDIGI (cakisma) durumlar
    hicbir tarafi otomatik ezmeden ayri bir "CAKISMALAR" sayfasina yazilir.

Kurulum ve kullanim icin README.md dosyasina bakin.
"""

from __future__ import annotations

import json
import logging
import re
import shutil
import sys
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from logging.handlers import RotatingFileHandler
from pathlib import Path
from typing import Optional

SCRIPT_DIR = Path(__file__).resolve().parent
CONFIG_PATH = SCRIPT_DIR / "config.json"
DEFAULT_STATE_PATH = SCRIPT_DIR / "sync_state.json"
DEFAULT_LOG_PATH = SCRIPT_DIR / "logs" / "sync.log"
DEFAULT_BACKUP_DIR = SCRIPT_DIR / "backups"

# Excel sayfasindaki sutun sirasi (1-indeksli). SENKRON_ID sutunu bu script
# tarafindan otomatik eklenir/kullanilir; elle silinmemeli.
COL_TARIH = 1
COL_GRUP = 2
COL_KUPE = 3
COL_TESHIS = 4
COL_TEDAVI = 5
COL_SENKRON_ID = 6

HEADERS = ["TARİH", "GRUP", "KÜPE", "TEŞHIS", "TEDAVİ", "SENKRON_ID"]

DAY_PREFIX_RE = re.compile(r"^\s*(\d+)\s*-\s*(.*)$", re.DOTALL)

CONFLICTS_SHEET_NAME = "CAKISMALAR (kontrol edin)"


# --------------------------------------------------------------------------
# Yapilandirma
# --------------------------------------------------------------------------


@dataclass
class Config:
    excel_path: Path
    sheet_name: str
    supabase_url: str
    supabase_anon_key: str
    sync_email: str
    sync_password: str
    sync_window_days: int = 30
    dry_run: bool = False
    state_path: Path = DEFAULT_STATE_PATH
    log_path: Path = DEFAULT_LOG_PATH
    backup_dir: Path = DEFAULT_BACKUP_DIR

    @staticmethod
    def load(path: Path = CONFIG_PATH) -> "Config":
        if not path.exists():
            raise SystemExit(
                f"Yapilandirma dosyasi bulunamadi: {path}\n"
                f"config.example.json dosyasini config.json olarak kopyalayip "
                f"kendi bilgilerinizi girin."
            )
        raw = json.loads(path.read_text(encoding="utf-8"))
        return Config(
            excel_path=Path(raw["excel_path"]),
            sheet_name=raw.get("sheet_name", "BUZAĞI TEDAVİ 2025"),
            supabase_url=raw["supabase_url"],
            supabase_anon_key=raw["supabase_anon_key"],
            sync_email=raw["sync_email"],
            sync_password=raw["sync_password"],
            sync_window_days=int(raw.get("sync_window_days", 30)),
            dry_run=bool(raw.get("dry_run", False)),
            state_path=Path(raw.get("state_path", DEFAULT_STATE_PATH)),
            log_path=Path(raw.get("log_path", DEFAULT_LOG_PATH)),
            backup_dir=Path(raw.get("backup_dir", DEFAULT_BACKUP_DIR)),
        )


def setup_logging(log_path: Path) -> logging.Logger:
    log_path.parent.mkdir(parents=True, exist_ok=True)
    logger = logging.getLogger("excel_sync")
    logger.setLevel(logging.INFO)
    logger.handlers.clear()  # tekrar cagrilirsa (orn. testlerde) handler'lar birikmesin
    fmt = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s", "%Y-%m-%d %H:%M:%S")

    file_handler = RotatingFileHandler(log_path, maxBytes=2_000_000, backupCount=5, encoding="utf-8")
    file_handler.setFormatter(fmt)
    logger.addHandler(file_handler)

    console_handler = logging.StreamHandler(sys.stdout)
    console_handler.setFormatter(fmt)
    logger.addHandler(console_handler)

    return logger


# --------------------------------------------------------------------------
# Excel tarafi
# --------------------------------------------------------------------------


@dataclass
class ExcelRow:
    row_number: int
    treatment_date: date
    group_raw: str
    ear_tag: str
    diagnosis: str
    protocol_day: Optional[int]
    description: str
    site_id: Optional[str]


def normalize_ear_tag(value) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, int):
        return str(value)
    text = str(value).strip()
    return text or None


def split_day_prefix(text: str) -> tuple[Optional[int], str]:
    """'5-REPTOPEN-...' -> (5, 'REPTOPEN-...'). Onek yoksa (None, text)."""
    if not text:
        return None, ""
    m = DAY_PREFIX_RE.match(text)
    if m:
        return int(m.group(1)), m.group(2).strip()
    return None, text.strip()


def join_day_prefix(protocol_day: Optional[int], description: str) -> str:
    if protocol_day is not None:
        return f"{protocol_day}-{description}"
    return description


def ensure_excel_writable(cfg: Config) -> None:
    """Supabase'e herhangi bir yazma islemi yapmadan ONCE dosyanin gercekten
    yazilabilir oldugunu dogrular (orn. Excel'de acik degil). Boylece site
    tarafinda degisiklik yapip da sonda Excel'e kaydedememe riski en aza iner."""
    try:
        with open(cfg.excel_path, "r+b"):
            pass
    except PermissionError as exc:
        raise SystemExit(
            f"Excel dosyasi acik veya kilitli, senkronizasyon atlandi: {exc}\n"
            f"Lutfen '{cfg.excel_path.name}' dosyasini kapatip tekrar deneyin."
        ) from exc


def load_excel(cfg: Config, logger: logging.Logger):
    import openpyxl

    if not cfg.excel_path.exists():
        raise SystemExit(f"Excel dosyasi bulunamadi: {cfg.excel_path}")

    try:
        wb = openpyxl.load_workbook(cfg.excel_path)
    except PermissionError as exc:
        raise SystemExit(
            f"Excel dosyasi acik veya kilitli, senkronizasyon atlandi: {exc}\n"
            f"Lutfen '{cfg.excel_path.name}' dosyasini kapatip tekrar deneyin."
        ) from exc

    if cfg.sheet_name not in wb.sheetnames:
        raise SystemExit(
            f"'{cfg.sheet_name}' adinda bir sayfa bulunamadi. Mevcut sayfalar: {wb.sheetnames}"
        )
    ws = wb[cfg.sheet_name]

    # SENKRON_ID sutunu yoksa ekle (sadece ilk calistirmada).
    header_row = [ws.cell(row=1, column=c).value for c in range(1, 7)]
    if header_row[COL_SENKRON_ID - 1] != "SENKRON_ID":
        ws.cell(row=1, column=COL_SENKRON_ID, value="SENKRON_ID")
        logger.info("SENKRON_ID sutunu eklendi (F sutunu).")

    return wb, ws


# Kullanicinin tarih sutununda kullandigi bicim: saat YOK, nokta ile
# ayrilmis (orn. "4.08.2026"). Sayfadan otomatik "tahmin etmeye"
# calismak (onceki surum) kararsizdi - onceki (hatali) calistirmalardan
# kalan hucreler yanlislikla ornek alinabiliyordu. Sabit deger daha
# guvenilir.
DATE_FORMAT = "d.mm.yyyy"


def detect_reference_style(ws) -> dict[int, dict]:
    """Yeni eklenen bir satirin yazi tipi/hizalama/kenarlik gibi gorunumu,
    kullanicinin elle girdigi satirlarla AYNI olsun diye: mevcut, gercek
    veri iceren bir satirin her sutunundaki hucre bicimini kopyalamak
    icin orneklerini toplar."""
    import copy

    for r in range(2, min(ws.max_row, 500) + 1):
        kupe_cell = ws.cell(row=r, column=COL_KUPE)
        if kupe_cell.value in (None, ""):
            continue
        styles = {}
        for col in (COL_TARIH, COL_GRUP, COL_KUPE, COL_TESHIS, COL_TEDAVI):
            src = ws.cell(row=r, column=col)
            styles[col] = {
                "font": copy.copy(src.font),
                "alignment": copy.copy(src.alignment),
                "border": copy.copy(src.border),
                "fill": copy.copy(src.fill),
            }
        return styles
    return {}


def apply_reference_style(ws, row_number: int, styles: dict[int, dict]) -> None:
    for col, style in styles.items():
        cell = ws.cell(row=row_number, column=col)
        cell.font = style["font"]
        cell.alignment = style["alignment"]
        cell.border = style["border"]
        cell.fill = style["fill"]


def read_excel_rows(ws, since: date) -> list[ExcelRow]:
    rows: list[ExcelRow] = []
    for r in range(2, ws.max_row + 1):
        raw_date = ws.cell(row=r, column=COL_TARIH).value
        raw_ear_tag = ws.cell(row=r, column=COL_KUPE).value
        if raw_date is None or raw_ear_tag is None:
            continue
        if isinstance(raw_date, datetime):
            treatment_date = raw_date.date()
        elif isinstance(raw_date, date):
            treatment_date = raw_date
        else:
            continue  # bozuk/tarih olmayan hucre - dokunma
        if treatment_date < since:
            continue

        ear_tag = normalize_ear_tag(raw_ear_tag)
        if not ear_tag:
            continue

        group_raw = str(ws.cell(row=r, column=COL_GRUP).value or "").strip()
        diagnosis = str(ws.cell(row=r, column=COL_TESHIS).value or "").strip()
        treatment_text = str(ws.cell(row=r, column=COL_TEDAVI).value or "").strip()
        protocol_day, description = split_day_prefix(treatment_text)
        site_id_raw = ws.cell(row=r, column=COL_SENKRON_ID).value
        site_id = str(site_id_raw).strip() if site_id_raw else None

        rows.append(
            ExcelRow(
                row_number=r,
                treatment_date=treatment_date,
                group_raw=group_raw,
                ear_tag=ear_tag,
                diagnosis=diagnosis,
                protocol_day=protocol_day,
                description=description,
                site_id=site_id,
            )
        )
    return rows


def write_excel_row(ws, row: ExcelRow) -> None:
    # Saat YOK - sadece tarih (datetime.min.time() ile gece yarisina
    # sabitleniyor, DATE_FORMAT da saat kismini hic gostermiyor).
    date_cell = ws.cell(row=row.row_number, column=COL_TARIH, value=datetime.combine(row.treatment_date, datetime.min.time()))
    date_cell.number_format = DATE_FORMAT
    ws.cell(row=row.row_number, column=COL_GRUP, value=row.group_raw)
    ws.cell(row=row.row_number, column=COL_KUPE, value=int(row.ear_tag) if row.ear_tag.isdigit() else row.ear_tag)
    ws.cell(row=row.row_number, column=COL_TESHIS, value=row.diagnosis)
    ws.cell(row=row.row_number, column=COL_TEDAVI, value=join_day_prefix(row.protocol_day, row.description))
    ws.cell(row=row.row_number, column=COL_SENKRON_ID, value=row.site_id)


def append_excel_row(ws, row: ExcelRow, reference_styles: Optional[dict[int, dict]] = None) -> int:
    new_row_number = ws.max_row + 1
    row.row_number = new_row_number
    write_excel_row(ws, row)
    if reference_styles:
        apply_reference_style(ws, new_row_number, reference_styles)
    return new_row_number


def write_conflicts_sheet(wb, conflicts: list[dict]) -> None:
    if CONFLICTS_SHEET_NAME in wb.sheetnames:
        del wb[CONFLICTS_SHEET_NAME]
    if not conflicts:
        return
    ws = wb.create_sheet(CONFLICTS_SHEET_NAME)
    ws.append(["TESPIT ZAMANI", "KÜPE", "TARİH", "AÇIKLAMA", "EXCEL DEĞERİ", "SİTE DEĞERİ", "SENKRON_ID"])
    for c in conflicts:
        ws.append(
            [
                c["detected_at"],
                c["ear_tag"],
                c["date"],
                c["reason"],
                c["excel_value"],
                c["site_value"],
                c.get("site_id", ""),
            ]
        )


# --------------------------------------------------------------------------
# Supabase tarafi
# --------------------------------------------------------------------------


@dataclass
class SiteTreatment:
    id: str
    animal_id: str
    ear_tag: str
    treatment_date: date
    diagnosis: str
    protocol_day: Optional[int]
    description: str


def connect(cfg: Config, logger: logging.Logger) -> tuple:
    from supabase import create_client

    client = create_client(cfg.supabase_url, cfg.supabase_anon_key)
    try:
        auth_response = client.auth.sign_in_with_password(
            {"email": cfg.sync_email, "password": cfg.sync_password}
        )
    except Exception as exc:  # noqa: BLE001 - kullaniciya net mesaj vermek icin genis yakala
        raise SystemExit(
            f"Siteye giris yapilamadi ({exc}). sync_email/sync_password dogru mu, "
            f"internet baglantisi var mi kontrol edin."
        ) from exc
    if not auth_response.user:
        raise SystemExit("Siteye giris yapilamadi: kullanici bilgisi donmedi.")
    logger.info("Siteye '%s' olarak giris yapildi.", cfg.sync_email)
    return client, auth_response.user.id


PAGE_SIZE = 1000


def fetch_all_pages(query_fn) -> list:
    """PostgREST istekleri varsayilan olarak en fazla 1000 satir dondurur;
    daha buyuk tablolarda (orn. yillar icinde biriken hayvanlar) sayfa sayfa
    cekmezsek son satirlar sessizce eksik kalir ve "zaten var" kayitlar
    yeniden olusturulmaya calisilip cakisma hatasi verir."""
    out: list = []
    offset = 0
    while True:
        page = query_fn(offset, offset + PAGE_SIZE - 1)
        out.extend(page)
        if len(page) < PAGE_SIZE:
            return out
        offset += PAGE_SIZE


def fetch_animals_by_ear_tag(client) -> dict[str, str]:
    rows = fetch_all_pages(
        lambda start, end: client.table("animals").select("id,ear_tag").range(start, end).execute().data
    )
    return {row["ear_tag"]: row["id"] for row in rows}


def fetch_site_treatments(client, since: date) -> list[SiteTreatment]:
    rows = fetch_all_pages(
        lambda start, end: client.table("calf_treatments")
        .select("id,animal_id,treatment_date,diagnosis,protocol_day,description,animals(ear_tag)")
        .gte("treatment_date", since.isoformat())
        .range(start, end)
        .execute()
        .data
    )
    out = []
    for row in rows:
        animal = row.get("animals") or {}
        ear_tag = animal.get("ear_tag") if isinstance(animal, dict) else None
        if not ear_tag:
            continue
        out.append(
            SiteTreatment(
                id=row["id"],
                animal_id=row["animal_id"],
                ear_tag=ear_tag,
                treatment_date=date.fromisoformat(row["treatment_date"]),
                diagnosis=row.get("diagnosis") or "",
                protocol_day=row.get("protocol_day"),
                description=row.get("description") or "",
            )
        )
    return out


def get_or_create_animal(client, logger: logging.Logger, ear_tag: str, animal_cache: dict[str, str], sync_profile_id: str) -> str:
    if ear_tag in animal_cache:
        return animal_cache[ear_tag]
    try:
        result = client.table("animals").insert(
            {
                "ear_tag": ear_tag,
                "gender": None,
                "status": "aktif",
                "created_by": sync_profile_id,
            }
        ).execute()
        animal_id = result.data[0]["id"]
        animal_cache[ear_tag] = animal_id
        logger.info("Yeni hayvan olusturuldu (kupe %s, Excel'den).", ear_tag)
        return animal_id
    except Exception as exc:  # noqa: BLE001
        # Bu kupe no sitede baska bir yerden (site arayuzu, es zamanli
        # calisma) zaten olusturulmus olabilir - hata yerine var olani bul.
        if getattr(exc, "code", None) == "23505" or "duplicate key" in str(exc):
            existing = client.table("animals").select("id").eq("ear_tag", ear_tag).execute()
            if existing.data:
                animal_id = existing.data[0]["id"]
                animal_cache[ear_tag] = animal_id
                logger.info("Kupe %s sitede zaten kayitliymis, tedavi ona baglaniyor.", ear_tag)
                return animal_id
        raise


def create_site_treatment(client, sync_profile_id: str, animal_id: str, row: ExcelRow) -> str:
    payload = {
        "animal_id": animal_id,
        "treatment_date": row.treatment_date.isoformat(),
        "diagnosis": row.diagnosis or None,
        "protocol_day": row.protocol_day,
        "description": row.description or row.diagnosis or "-",
        "created_by": sync_profile_id,
        "course_id": None,
    }
    result = client.table("calf_treatments").insert(payload).execute()
    return result.data[0]["id"]


def update_site_treatment(client, site_id: str, row: ExcelRow) -> None:
    payload = {
        "treatment_date": row.treatment_date.isoformat(),
        "diagnosis": row.diagnosis or None,
        "protocol_day": row.protocol_day,
        "description": row.description or row.diagnosis or "-",
    }
    client.table("calf_treatments").update(payload).eq("id", site_id).execute()


def fetch_treatment_by_id(client, site_id: str) -> Optional[dict]:
    result = client.table("calf_treatments").select("id").eq("id", site_id).execute()
    return result.data[0] if result.data else None


def fetch_current_location(client, animal_id: str) -> str:
    for structure, label_fn in (
        ("buzagilik", lambda g: "BUZAĞILIK"),
        ("iglo", lambda g: f"{g + 1}.İGLO"),
    ):
        result = (
            client.table("calf_housing_slots")
            .select("group_index")
            .eq("structure", structure)
            .eq("animal_id", animal_id)
            .limit(1)
            .execute()
        )
        if result.data:
            return label_fn(result.data[0]["group_index"])
    # Site sadece hayvanin SU ANKI kulube atamasini tutar, gecmis
    # tedavinin hangi odada yapildigini degil - hayvan artik hicbir
    # kulubede degilse (sutten kesilmis, satilmis, vb.) geriye donuk
    # gercek konum bilinemez. Sitenin kendi arayuzunde de ayni durum
    # icin kullanilan ifade.
    return "Barınakta değil"


# --------------------------------------------------------------------------
# Senkronizasyon durumu (son bilinen esitlenmis deger, cakisma tespiti icin)
# --------------------------------------------------------------------------


def snapshot_of(ear_tag: str, treatment_date: date, diagnosis: str, protocol_day: Optional[int], description: str) -> dict:
    return {
        "ear_tag": ear_tag,
        "date": treatment_date.isoformat(),
        "diagnosis": diagnosis.strip(),
        "protocol_day": protocol_day,
        "description": description.strip(),
    }


def load_state(path: Path) -> dict:
    if path.exists():
        return json.loads(path.read_text(encoding="utf-8"))
    return {}


def save_state(path: Path, state: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8")


def natural_match_key(ear_tag: str, treatment_date: date, description: str) -> tuple:
    return (ear_tag, treatment_date.isoformat(), description.strip().casefold())


# --------------------------------------------------------------------------
# Ana senkronizasyon
# --------------------------------------------------------------------------


@dataclass
class SyncStats:
    pushed_to_site: int = 0
    pulled_to_excel: int = 0
    linked: int = 0
    updated_on_site: int = 0
    conflicts: list = field(default_factory=list)
    errors: list = field(default_factory=list)


def run_sync(cfg: Config, logger: logging.Logger) -> SyncStats:
    stats = SyncStats()
    since = date.today() - timedelta(days=cfg.sync_window_days)
    now_str = datetime.now().isoformat(timespec="seconds")

    logger.info("Senkronizasyon basliyor. Pencere: %s -> bugun.", since.isoformat())

    if not cfg.dry_run:
        cfg.backup_dir.mkdir(parents=True, exist_ok=True)
        backup_name = f"{cfg.excel_path.stem}_{datetime.now():%Y%m%d_%H%M%S}{cfg.excel_path.suffix}"
        shutil.copy2(cfg.excel_path, cfg.backup_dir / backup_name)
        logger.info("Yedek alindi: %s", backup_name)

    wb, ws = load_excel(cfg, logger)
    reference_styles = detect_reference_style(ws)
    excel_rows = read_excel_rows(ws, since)
    logger.info("Excel'de pencere icinde %d satir bulundu.", len(excel_rows))

    if not cfg.dry_run:
        ensure_excel_writable(cfg)

    client, sync_profile_id = connect(cfg, logger)

    animal_cache = fetch_animals_by_ear_tag(client)
    site_treatments = fetch_site_treatments(client, since)
    site_by_id = {t.id: t for t in site_treatments}
    logger.info("Sitede pencere icinde %d tedavi kaydi bulundu.", len(site_treatments))

    state = load_state(cfg.state_path)
    conflicts: list[dict] = []

    excel_by_natural_key: dict[tuple, ExcelRow] = {}
    linked_site_ids: set[str] = set()

    # 1) SENKRON_ID'si olan Excel satirlari: karsilastir, gerekirse siteye
    #    guncelleme gonder, ya da siteden silinmisse isaretle.
    for row in excel_rows:
        if not row.site_id:
            excel_by_natural_key[natural_match_key(row.ear_tag, row.treatment_date, row.description)] = row
            continue

        linked_site_ids.add(row.site_id)
        current_snapshot = snapshot_of(row.ear_tag, row.treatment_date, row.diagnosis, row.protocol_day, row.description)
        last_snapshot = state.get(row.site_id)
        site_row = site_by_id.get(row.site_id)

        if site_row is None:
            # Pencere disinda mi yoksa gercekten silinmis mi diye kontrol et.
            exists = fetch_treatment_by_id(client, row.site_id)
            if exists is None:
                conflicts.append(
                    {
                        "detected_at": now_str,
                        "ear_tag": row.ear_tag,
                        "date": row.treatment_date.isoformat(),
                        "reason": "Sitede silinmis (Geri Al ile), Excel'de hala duruyor",
                        "excel_value": join_day_prefix(row.protocol_day, row.description),
                        "site_value": "(silindi)",
                        "site_id": row.site_id,
                    }
                )
            continue

        if last_snapshot is None:
            # Ilk kez goruyoruz (durum dosyasi kayboldu/ilk calistirma) -
            # Excel'i temel al, siteyle karsilastirip farkli ise guncelle.
            last_snapshot = current_snapshot

        excel_changed = current_snapshot != last_snapshot
        site_snapshot = snapshot_of(site_row.ear_tag, site_row.treatment_date, site_row.diagnosis, site_row.protocol_day, site_row.description)
        site_changed = site_snapshot != last_snapshot

        if excel_changed and site_changed and current_snapshot != site_snapshot:
            conflicts.append(
                {
                    "detected_at": now_str,
                    "ear_tag": row.ear_tag,
                    "date": row.treatment_date.isoformat(),
                    "reason": "Hem Excel'de hem sitede degismis",
                    "excel_value": join_day_prefix(row.protocol_day, row.description),
                    "site_value": join_day_prefix(site_row.protocol_day, site_row.description),
                    "site_id": row.site_id,
                }
            )
            continue

        if excel_changed:
            if not cfg.dry_run:
                update_site_treatment(client, row.site_id, row)
            state[row.site_id] = current_snapshot
            stats.updated_on_site += 1
            logger.info("Kupe %s (%s) siteye guncellendi (Excel'de degismisti).", row.ear_tag, row.treatment_date)
        elif site_changed:
            row.diagnosis = site_row.diagnosis
            row.protocol_day = site_row.protocol_day
            row.description = site_row.description
            if not cfg.dry_run:
                write_excel_row(ws, row)
            state[row.site_id] = site_snapshot
            stats.pulled_to_excel += 1

    # 2) SENKRON_ID'si olmayan Excel satirlari: siteye yeni kayit olarak ekle
    #    (once dogal anahtar ile zaten sitede var mi diye bak, varsa sadece
    #    baglantiyi kur).
    natural_key_to_site = {
        natural_match_key(t.ear_tag, t.treatment_date, t.description): t
        for t in site_treatments
        if t.id not in linked_site_ids
    }

    for row in excel_rows:
        if row.site_id:
            continue
        key = natural_match_key(row.ear_tag, row.treatment_date, row.description)
        existing = natural_key_to_site.get(key)
        if existing is not None:
            row.site_id = existing.id
            linked_site_ids.add(existing.id)
            if not cfg.dry_run:
                write_excel_row(ws, row)
            state[existing.id] = snapshot_of(row.ear_tag, row.treatment_date, row.diagnosis, row.protocol_day, row.description)
            stats.linked += 1
            continue

        try:
            animal_id = get_or_create_animal(client, logger, row.ear_tag, animal_cache, sync_profile_id) if not cfg.dry_run else animal_cache.get(row.ear_tag, "dry-run")
            new_id = create_site_treatment(client, sync_profile_id, animal_id, row) if not cfg.dry_run else "dry-run-id"
            row.site_id = new_id
            if not cfg.dry_run:
                write_excel_row(ws, row)
                state[new_id] = snapshot_of(row.ear_tag, row.treatment_date, row.diagnosis, row.protocol_day, row.description)
            stats.pushed_to_site += 1
            logger.info("Kupe %s (%s) siteye eklendi.", row.ear_tag, row.treatment_date)
        except Exception as exc:  # noqa: BLE001
            logger.error("Kupe %s (satir %d) siteye eklenemedi: %s", row.ear_tag, row.row_number, exc)
            stats.errors.append(f"Satir {row.row_number} (kupe {row.ear_tag}): {exc}")

    # 3) Sitede olup Excel'de hic gorunmeyen kayitlar: Excel'e yeni satir
    #    olarak ekle. Once dogal anahtarla eslesenleri baglayip devre disi
    #    birak, gercekten yeni olanlari ise TEK TEK eklemek yerine
    #    kupe numarasi + tedavi tarihine gore sirala ve OYLE ekle - aksi
    #    halde veritabanindan gelen sirali OLMAYAN sira ile satirlar
    #    "2-3-2-5" gibi karisik/kronolojik olmayan bir duzende Excel'e
    #    dusuyordu (kullanicinin elle girdigi satirlar her zaman kupeye
    #    gore gruplanip tarihe gore artan sirada duzenli).
    genuinely_new: list[SiteTreatment] = []
    for t in site_treatments:
        if t.id in linked_site_ids:
            continue
        key = natural_match_key(t.ear_tag, t.treatment_date, t.description)
        if key in excel_by_natural_key:
            existing_row = excel_by_natural_key[key]
            existing_row.site_id = t.id
            if not cfg.dry_run:
                write_excel_row(ws, existing_row)
                state[t.id] = snapshot_of(existing_row.ear_tag, existing_row.treatment_date, existing_row.diagnosis, existing_row.protocol_day, existing_row.description)
            stats.linked += 1
            continue
        genuinely_new.append(t)

    def sort_key(t: SiteTreatment):
        ear_tag = t.ear_tag or ""
        ear_sort = (0, int(ear_tag)) if ear_tag.isdigit() else (1, ear_tag)
        return (ear_sort, t.treatment_date, t.protocol_day if t.protocol_day is not None else 0)

    genuinely_new.sort(key=sort_key)

    for t in genuinely_new:
        group_raw = fetch_current_location(client, t.animal_id) if not cfg.dry_run else "-"
        new_row = ExcelRow(
            row_number=-1,
            treatment_date=t.treatment_date,
            group_raw=group_raw,
            ear_tag=t.ear_tag,
            diagnosis=t.diagnosis,
            protocol_day=t.protocol_day,
            description=t.description,
            site_id=t.id,
        )
        if not cfg.dry_run:
            append_excel_row(ws, new_row, reference_styles)
            state[t.id] = snapshot_of(t.ear_tag, t.treatment_date, t.diagnosis, t.protocol_day, t.description)
        stats.pulled_to_excel += 1
        logger.info("Kupe %s (%s) Excel'e eklendi (siteden).", t.ear_tag, t.treatment_date)

    stats.conflicts = conflicts

    if not cfg.dry_run:
        write_conflicts_sheet(wb, conflicts)
        try:
            wb.save(cfg.excel_path)
        except PermissionError as exc:
            raise SystemExit(
                f"Excel dosyasina kaydedilemedi (dosya tekrar acilmis olabilir): {exc}\n"
                f"Site tarafinda yapilan degisiklikler bir sonraki calistirmada "
                f"otomatik olarak Excel'e yansiyacak."
            ) from exc
        save_state(cfg.state_path, state)

    return stats


def main() -> int:
    cfg = Config.load()
    logger = setup_logging(cfg.log_path)
    try:
        stats = run_sync(cfg, logger)
    except SystemExit as exc:
        logger.error(str(exc))
        return 1
    except Exception:  # noqa: BLE001
        logger.exception("Beklenmeyen hata, senkronizasyon durduruldu.")
        return 1

    logger.info(
        "Tamamlandi. Siteye eklenen: %d, guncellenen: %d, Excel'e eklenen: %d, "
        "baglanan: %d, cakisma: %d, hata: %d%s",
        stats.pushed_to_site,
        stats.updated_on_site,
        stats.pulled_to_excel,
        stats.linked,
        len(stats.conflicts),
        len(stats.errors),
        " [DRY RUN - hicbir sey yazilmadi]" if cfg.dry_run else "",
    )
    if stats.conflicts:
        logger.warning(
            "%d cakisma bulundu, '%s' sayfasini kontrol edin.", len(stats.conflicts), CONFLICTS_SHEET_NAME
        )
    return 0 if not stats.errors else 2


if __name__ == "__main__":
    raise SystemExit(main())
