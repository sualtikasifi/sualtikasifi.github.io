"""
sync.py icin regresyon testleri.

Gercek Supabase'e baglanmadan calisir: Supabase Python istemcisinin
kullandigimiz kucuk bir kismini (auth + table().select/insert/update/eq/
gte/lte/limit/execute) bellek-ici sahte bir veritabaniyla taklit eder, ve
gercek bir gecici .xlsx dosyasi uzerinde okuma/yazma yapar.

Calistirmak icin:
    python3 test_sync.py
"""

from __future__ import annotations

import sys
import tempfile
import uuid
from datetime import date, datetime, timedelta
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

import openpyxl

import sync as S


# --------------------------------------------------------------------------
# Sahte Supabase istemcisi
# --------------------------------------------------------------------------


class FakeResult:
    def __init__(self, data):
        self.data = data


class FakeAuthUser:
    def __init__(self, user_id):
        self.id = user_id


class FakeAuthResponse:
    def __init__(self, user_id):
        self.user = FakeAuthUser(user_id)


class FakeAuth:
    def __init__(self, profile_id: str, valid_password: str):
        self.profile_id = profile_id
        self.valid_password = valid_password

    def sign_in_with_password(self, creds):
        if creds.get("password") != self.valid_password:
            raise RuntimeError("gecersiz sifre")
        return FakeAuthResponse(self.profile_id)


class FakeAPIError(Exception):
    """postgrest.exceptions.APIError'in sync.py'nin kontrol ettigi kismini
    (str(exc) icinde mesaj, .code niteligi) taklit eder."""

    def __init__(self, message: str, code: str):
        super().__init__({"message": message, "code": code, "hint": None, "details": None})
        self.code = code


class FakeQuery:
    def __init__(self, db, table_name):
        self.db = db
        self.table_name = table_name
        self._rows = None
        self._select_cols = None
        self._filters = []
        self._limit = None
        self._range = None
        self._mode = None
        self._payload = None

    # --- filtre/sec zinciri ---
    def select(self, cols):
        self._select_cols = cols
        self._mode = "select"
        return self

    def eq(self, col, val):
        self._filters.append(("eq", col, val))
        return self

    def gte(self, col, val):
        self._filters.append(("gte", col, val))
        return self

    def lte(self, col, val):
        self._filters.append(("lte", col, val))
        return self

    def limit(self, n):
        self._limit = n
        return self

    def range(self, start, end):
        self._range = (start, end)
        return self

    def insert(self, payload):
        self._mode = "insert"
        self._payload = payload
        return self

    def update(self, payload):
        self._mode = "update"
        self._payload = payload
        return self

    # --- calistir ---
    def execute(self):
        table = self.db.setdefault(self.table_name, [])

        if self._mode == "insert":
            row = dict(self._payload)
            if self.table_name == "animals" and any(a["ear_tag"] == row.get("ear_tag") for a in table):
                raise FakeAPIError(
                    'duplicate key value violates unique constraint "animals_ear_tag_key"', code="23505"
                )
            row.setdefault("id", str(uuid.uuid4()))
            row.setdefault("created_at", datetime.now().isoformat())
            table.append(row)
            return FakeResult([dict(row)])

        if self._mode == "update":
            matched = self._apply_filters(table)
            for row in matched:
                row.update(self._payload)
            return FakeResult([dict(r) for r in matched])

        # select
        matched = self._apply_filters(table)
        if self._range is not None:
            start, end = self._range
            matched = matched[start : end + 1]
        if self._limit is not None:
            matched = matched[: self._limit]

        out = []
        for row in matched:
            projected = dict(row)
            if self._select_cols and "animals(ear_tag)" in self._select_cols:
                animal = next((a for a in self.db.get("animals", []) if a["id"] == row.get("animal_id")), None)
                projected["animals"] = {"ear_tag": animal["ear_tag"]} if animal else None
            out.append(projected)
        return FakeResult(out)

    def _apply_filters(self, table):
        matched = table
        for op, col, val in self._filters:
            if op == "eq":
                matched = [r for r in matched if r.get(col) == val]
            elif op == "gte":
                matched = [r for r in matched if str(r.get(col, "")) >= str(val)]
            elif op == "lte":
                matched = [r for r in matched if str(r.get(col, "")) <= str(val)]
        return matched


class FakeClient:
    def __init__(self, profile_id: str, valid_password: str):
        self.db: dict[str, list] = {"animals": [], "calf_treatments": [], "calf_housing_slots": []}
        self.auth = FakeAuth(profile_id, valid_password)

    def table(self, name):
        return FakeQuery(self.db, name)


# --------------------------------------------------------------------------
# Test yardimcilari
# --------------------------------------------------------------------------

PASSED = 0
FAILED = 0


def check(name: str, condition: bool, detail: str = ""):
    global PASSED, FAILED
    if condition:
        PASSED += 1
        print(f"  OK  {name}")
    else:
        FAILED += 1
        print(f"  FAIL {name}  {detail}")


def make_workbook(tmp_path: Path, rows: list[list]) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BUZAĞI TEDAVİ 2025"
    ws.append(["TARİH", "GRUP", "KÜPE", "TEŞHIS", "TEDAVİ"])
    for row in rows:
        ws.append(row)
    path = tmp_path / "test.xlsx"
    wb.save(path)
    return path


def make_config(tmp_path: Path, excel_path: Path, sync_window_days=30) -> S.Config:
    return S.Config(
        excel_path=excel_path,
        sheet_name="BUZAĞI TEDAVİ 2025",
        supabase_url="http://fake",
        supabase_anon_key="fake",
        sync_email="sync@example.com",
        sync_password="secret",
        sync_window_days=sync_window_days,
        dry_run=False,
        state_path=tmp_path / "state.json",
        log_path=tmp_path / "logs" / "sync.log",
        backup_dir=tmp_path / "backups",
    )


class Ctx:
    """run_sync tek bir connect() cagrisi bekliyor; testte gercek supabase
    yerine FakeClient donmesi icin sync.connect'i gecici olarak degistiririz."""

    def __init__(self, client: FakeClient, profile_id: str):
        self.client = client
        self.profile_id = profile_id

    def __enter__(self):
        self._orig = S.connect
        S.connect = lambda cfg, logger: (self.client, self.profile_id)
        return self

    def __exit__(self, *a):
        S.connect = self._orig


def read_sheet(path: Path):
    wb = openpyxl.load_workbook(path)
    ws = wb["BUZAĞI TEDAVİ 2025"]
    rows = []
    for r in range(2, ws.max_row + 1):
        rows.append([ws.cell(row=r, column=c).value for c in range(1, 7)])
    return rows


# --------------------------------------------------------------------------
# Testler
# --------------------------------------------------------------------------


def test_day_prefix_parsing():
    day, desc = S.split_day_prefix("5-REPTOPEN-PRİMOXAL-C VİT-B VİT-")
    check("split_day_prefix gun numarasini ayirir", day == 5 and desc == "REPTOPEN-PRİMOXAL-C VİT-B VİT-", f"got ({day!r}, {desc!r})")

    day, desc = S.split_day_prefix("BAYTRİL+ACTİMİSİN+B VİT")
    check("split_day_prefix onek yoksa None doner", day is None and desc == "BAYTRİL+ACTİMİSİN+B VİT", f"got ({day!r}, {desc!r})")

    check("join_day_prefix geri birlestirir", S.join_day_prefix(5, "X-Y") == "5-X-Y")
    check("join_day_prefix gun yoksa duz metin", S.join_day_prefix(None, "X-Y") == "X-Y")


def test_ear_tag_normalization():
    check("int kupe no", S.normalize_ear_tag(4262) == "4262")
    check("float kupe no (excel genelde float verir)", S.normalize_ear_tag(4262.0) == "4262")
    check("string kupe no", S.normalize_ear_tag("TR-70") == "TR-70")
    check("bos kupe no", S.normalize_ear_tag(None) is None)


def test_new_excel_row_pushed_to_site():
    with tempfile.TemporaryDirectory() as td:
        tmp = Path(td)
        today = date.today()
        excel_path = make_workbook(tmp, [[datetime.combine(today, datetime.min.time()), "BUZAĞILIK", 9001, "İSHAL", "3-DİGEST-B VİT"]])
        cfg = make_config(tmp, excel_path)

        client = FakeClient("sync-profile-id", "secret")
        with Ctx(client, "sync-profile-id"):
            stats = S.run_sync(cfg, S.setup_logging(cfg.log_path))

        check("yeni excel satiri siteye eklendi", stats.pushed_to_site == 1, str(stats))
        check("yeni hayvan olusturuldu", any(a["ear_tag"] == "9001" for a in client.db["animals"]))
        treatment = client.db["calf_treatments"][0]
        check("teshis dogru aktarildi", treatment["diagnosis"] == "İSHAL")
        check("gun numarasi ayrildi", treatment["protocol_day"] == 3)
        check("aciklama gun numarasiz", treatment["description"] == "DİGEST-B VİT")

        rows = read_sheet(excel_path)
        check("excel'e SENKRON_ID yazildi", rows[0][5] not in (None, ""), rows)


def test_new_site_treatment_pulled_to_excel():
    with tempfile.TemporaryDirectory() as td:
        tmp = Path(td)
        excel_path = make_workbook(tmp, [])
        cfg = make_config(tmp, excel_path)

        client = FakeClient("sync-profile-id", "secret")
        animal_id = str(uuid.uuid4())
        client.db["animals"].append({"id": animal_id, "ear_tag": "500"})
        client.db["calf_treatments"].append(
            {
                "id": str(uuid.uuid4()),
                "animal_id": animal_id,
                "treatment_date": date.today().isoformat(),
                "diagnosis": "PNÖMONİ",
                "protocol_day": 1,
                "description": "SELECTAN-B VİT",
                "created_by": "someone-else",
            }
        )

        with Ctx(client, "sync-profile-id"):
            stats = S.run_sync(cfg, S.setup_logging(cfg.log_path))

        check("site kaydi excel'e eklendi", stats.pulled_to_excel == 1, str(stats))
        rows = read_sheet(excel_path)
        # Kupe no orijinal dosyadaki gibi sayi olarak yazilir (metin degil).
        check("excel satiri kupe eslesiyor", rows[0][2] == 500, rows)
        check("excel satiri tedavi metni gun+aciklama", rows[0][4] == "1-SELECTAN-B VİT", rows)


def test_idempotent_second_run_is_noop():
    with tempfile.TemporaryDirectory() as td:
        tmp = Path(td)
        today = date.today()
        excel_path = make_workbook(tmp, [[datetime.combine(today, datetime.min.time()), "BUZAĞILIK", 42, "İSHAL", "1-X"]])
        cfg = make_config(tmp, excel_path)
        client = FakeClient("sync-profile-id", "secret")

        with Ctx(client, "sync-profile-id"):
            S.run_sync(cfg, S.setup_logging(cfg.log_path))
            stats2 = S.run_sync(cfg, S.setup_logging(cfg.log_path))

        check(
            "ikinci calistirma hicbir sey degistirmiyor",
            (stats2.pushed_to_site, stats2.pulled_to_excel, stats2.linked, stats2.updated_on_site, len(stats2.conflicts))
            == (0, 0, 0, 0, 0),
            str(stats2),
        )
        check("hala tek tedavi kaydi var (duplike yok)", len(client.db["calf_treatments"]) == 1)


def test_excel_edit_pushes_update():
    with tempfile.TemporaryDirectory() as td:
        tmp = Path(td)
        today = date.today()
        excel_path = make_workbook(tmp, [[datetime.combine(today, datetime.min.time()), "BUZAĞILIK", 77, "İSHAL", "1-ESKI"]])
        cfg = make_config(tmp, excel_path)
        client = FakeClient("sync-profile-id", "secret")

        with Ctx(client, "sync-profile-id"):
            S.run_sync(cfg, S.setup_logging(cfg.log_path))

            wb = openpyxl.load_workbook(excel_path)
            ws = wb["BUZAĞI TEDAVİ 2025"]
            ws.cell(row=2, column=S.COL_TEDAVI, value="1-YENİ İLAÇ")
            wb.save(excel_path)

            stats2 = S.run_sync(cfg, S.setup_logging(cfg.log_path))

        check("excel duzenlemesi siteye push edildi", stats2.updated_on_site == 1, str(stats2))
        check("site kaydi guncellendi", client.db["calf_treatments"][0]["description"] == "YENİ İLAÇ")


def test_conflict_when_both_sides_differ():
    with tempfile.TemporaryDirectory() as td:
        tmp = Path(td)
        today = date.today()
        excel_path = make_workbook(tmp, [[datetime.combine(today, datetime.min.time()), "BUZAĞILIK", 88, "İSHAL", "1-A"]])
        cfg = make_config(tmp, excel_path)
        client = FakeClient("sync-profile-id", "secret")

        with Ctx(client, "sync-profile-id"):
            S.run_sync(cfg, S.setup_logging(cfg.log_path))

            # Excel'de degistir.
            wb = openpyxl.load_workbook(excel_path)
            ws = wb["BUZAĞI TEDAVİ 2025"]
            ws.cell(row=2, column=S.COL_TEDAVI, value="1-EXCELDEN DEGISTI")
            wb.save(excel_path)

            # State dosyasini bozmadan, sitedeki kaydi da DOGRUDAN (script disinda,
            # yani "site tarafinda baska bir yerden degisti" senaryosunu taklit
            # ederek) degistir.
            client.db["calf_treatments"][0]["description"] = "SITEDEN DEGISTI"

            stats2 = S.run_sync(cfg, S.setup_logging(cfg.log_path))

        check("cakisma tespit edildi", len(stats2.conflicts) == 1, str(stats2.conflicts))
        check("hicbir taraf otomatik ezilmedi (site degeri hala kendi degeri)", client.db["calf_treatments"][0]["description"] == "SITEDEN DEGISTI")
        rows = read_sheet(excel_path)
        check("excel degeri de otomatik ezilmedi", rows[0][4] == "1-EXCELDEN DEGISTI", rows)


def test_site_deletion_flagged_not_silently_removed():
    with tempfile.TemporaryDirectory() as td:
        tmp = Path(td)
        today = date.today()
        excel_path = make_workbook(tmp, [[datetime.combine(today, datetime.min.time()), "BUZAĞILIK", 99, "İSHAL", "1-A"]])
        cfg = make_config(tmp, excel_path)
        client = FakeClient("sync-profile-id", "secret")

        with Ctx(client, "sync-profile-id"):
            S.run_sync(cfg, S.setup_logging(cfg.log_path))
            client.db["calf_treatments"].clear()  # "Geri Al" ile silindi
            stats2 = S.run_sync(cfg, S.setup_logging(cfg.log_path))

        check("silinen kayit cakisma olarak bildirildi", len(stats2.conflicts) == 1, str(stats2.conflicts))
        rows = read_sheet(excel_path)
        check("excel satiri silinmedi (guvenli varsayilan)", len(rows) == 1, rows)


def test_dry_run_does_not_write_anything():
    with tempfile.TemporaryDirectory() as td:
        tmp = Path(td)
        today = date.today()
        excel_path = make_workbook(tmp, [[datetime.combine(today, datetime.min.time()), "BUZAĞILIK", 11, "İSHAL", "1-A"]])
        original_bytes = excel_path.read_bytes()
        cfg = make_config(tmp, excel_path)
        cfg.dry_run = True
        client = FakeClient("sync-profile-id", "secret")

        with Ctx(client, "sync-profile-id"):
            stats = S.run_sync(cfg, S.setup_logging(cfg.log_path))

        check("dry run sirasinda siteye hicbir sey yazilmadi", len(client.db["calf_treatments"]) == 0, str(client.db))
        check("dry run sirasinda excel dosyasi degismedi", excel_path.read_bytes() == original_bytes)
        check("dry run yine de dogru sayiyor", stats.pushed_to_site == 1)


def test_sync_window_ignores_old_rows():
    with tempfile.TemporaryDirectory() as td:
        tmp = Path(td)
        old_date = date.today() - timedelta(days=400)
        excel_path = make_workbook(tmp, [[datetime.combine(old_date, datetime.min.time()), "BUZAĞILIK", 22, "ESKİ", "1-A"]])
        cfg = make_config(tmp, excel_path, sync_window_days=30)
        client = FakeClient("sync-profile-id", "secret")

        with Ctx(client, "sync-profile-id"):
            stats = S.run_sync(cfg, S.setup_logging(cfg.log_path))

        check("pencere disindaki eski satir dokunulmadan atlandi", stats.pushed_to_site == 0)
        check("eski satiya SENKRON_ID yazilmadi", len(client.db["calf_treatments"]) == 0)


def test_locked_file_fails_fast_without_touching_site():
    """ensure_excel_writable() dogrudan test edilir: root olarak calisan bir
    test ortaminda gercek dosya izinleri islemedigi icin (root her seyi
    yazabilir), PermissionError'i acik/kilitli dosya durumunu taklit etmek
    icin open() uzerinden simule ediyoruz - boylece Windows'ta Excel'in
    dosyayi kilitledigi durumla ayni kod yolu, ortamdan bagimsiz test edilir.
    """
    with tempfile.TemporaryDirectory() as td:
        tmp = Path(td)
        today = date.today()
        excel_path = make_workbook(tmp, [[datetime.combine(today, datetime.min.time()), "BUZAĞILIK", 33, "İSHAL", "1-A"]])
        cfg = make_config(tmp, excel_path)
        client = FakeClient("sync-profile-id", "secret")

        real_open = S.__builtins__["open"] if isinstance(S.__builtins__, dict) else S.__builtins__.open

        def fake_open(path, mode="r", *a, **kw):
            if str(path) == str(excel_path) and "r+b" in mode:
                raise PermissionError("dosya baska bir program tarafindan kullaniliyor")
            return real_open(path, mode, *a, **kw)

        S.open = fake_open
        try:
            with Ctx(client, "sync-profile-id"):
                try:
                    S.run_sync(cfg, S.setup_logging(cfg.log_path))
                    raised = False
                except SystemExit:
                    raised = True
        finally:
            del S.open

        check("kilitli dosyada SystemExit firlatildi", raised)
        check("siteye hicbir sey yazilmadi", len(client.db["calf_treatments"]) == 0 and len(client.db["animals"]) == 0, str(client.db))


def test_new_rows_match_existing_date_format():
    """Gercek olayin regresyon testi: yeni eklenen satirlarin tarih hucresi
    openpyxl'in varsayilan (genis) formatinda degil, kullanicinin sayfada
    zaten kullandigi formatta (orn. 4.08.2026) olmali - aksi halde sutun
    dar oldugunda Excel '#####' gosteriyordu."""
    with tempfile.TemporaryDirectory() as td:
        tmp = Path(td)
        today = date.today()
        excel_path = make_workbook(
            tmp, [[datetime.combine(today, datetime.min.time()), "BUZAĞILIK", 10, "İSHAL", "1-A"]]
        )
        wb = openpyxl.load_workbook(excel_path)
        ws = wb["BUZAĞI TEDAVİ 2025"]
        ws.cell(row=2, column=S.COL_TARIH).number_format = "d.mm.yyyy"
        wb.save(excel_path)

        cfg = make_config(tmp, excel_path)
        client = FakeClient("sync-profile-id", "secret")
        animal_id = str(uuid.uuid4())
        client.db["animals"].append({"id": animal_id, "ear_tag": "999"})
        client.db["calf_treatments"].append(
            {
                "id": str(uuid.uuid4()),
                "animal_id": animal_id,
                "treatment_date": today.isoformat(),
                "diagnosis": "PNÖMONİ",
                "protocol_day": 1,
                "description": "X",
                "created_by": "someone-else",
            }
        )

        with Ctx(client, "sync-profile-id"):
            S.run_sync(cfg, S.setup_logging(cfg.log_path))

        wb2 = openpyxl.load_workbook(excel_path)
        ws2 = wb2["BUZAĞI TEDAVİ 2025"]
        new_row_format = ws2.cell(row=3, column=S.COL_TARIH).number_format
        check("yeni satirin tarih formati eskiyle ayni", new_row_format == "d.mm.yyyy", new_row_format)


def test_animal_lookup_survives_over_1000_existing_animals():
    """Gercek olayin regresyon testi: PostgREST varsayilan olarak sayfa
    basina en fazla 1000 satir dondurur. animals tablosunda 1000'den fazla
    kayit varsa ve fetch_animals_by_ear_tag sayfalama yapmiyorsa, son
    satirlar sessizce cache'e girmiyor ve o hayvanlar icin "zaten var"
    olan bir kupe numarasi yeniden olusturulmaya calisilip cakisma
    hatasi aliniyordu."""
    with tempfile.TemporaryDirectory() as td:
        tmp = Path(td)
        today = date.today()
        # Excel'de, halihazirda sitede kayitli (ama 1000. satirdan SONRA
        # eklenmis, yani sayfalanmadan cekilirse gorulmeyecek) bir hayvana
        # ait tedavi satiri var.
        excel_path = make_workbook(
            tmp, [[datetime.combine(today, datetime.min.time()), "BUZAĞILIK", "9999", "İSHAL", "1-B VİT"]]
        )
        cfg = make_config(tmp, excel_path)
        client = FakeClient("sync-profile-id", "secret")

        for i in range(1200):
            client.db["animals"].append({"id": str(uuid.uuid4()), "ear_tag": str(i)})
        # 9999 numarali kupe, ilk 1000'in DISINDA - sayfalama olmadan
        # cache'e girmeyecek konumda.
        client.db["animals"].append({"id": str(uuid.uuid4()), "ear_tag": "9999"})

        with Ctx(client, "sync-profile-id"):
            stats = S.run_sync(cfg, S.setup_logging(cfg.log_path))

        check("1200+ hayvan arasinda dogru hayvan bulundu, cakisma hatasi yok", len(stats.errors) == 0, str(stats.errors))
        check("zaten var olan hayvan tekrar olusturulmadi (dogru sayida hayvan)", len(client.db["animals"]) == 1201, len(client.db["animals"]))
        check("tedavi dogru hayvana baglandi", stats.linked == 0 and stats.pushed_to_site == 1, str(stats))


def test_duplicate_key_falls_back_to_existing_animal():
    """Yukaridaki sayfalama duzeltmesi olmasa bile (orn. baska bir yol
    yuzunden ayni anda olusma gibi), duplicate key hatasi sessizce
    hata olarak loglanmak yerine var olan hayvani bulup kullanmali."""
    with tempfile.TemporaryDirectory() as td:
        tmp = Path(td)
        today = date.today()
        excel_path = make_workbook(
            tmp, [[datetime.combine(today, datetime.min.time()), "BUZAĞILIK", "555", "İSHAL", "1-A"]]
        )
        cfg = make_config(tmp, excel_path)
        client = FakeClient("sync-profile-id", "secret")
        existing_id = str(uuid.uuid4())
        client.db["animals"].append({"id": existing_id, "ear_tag": "555"})

        # fetch_animals_by_ear_tag'in bu hayvani KACIRDIGINI simule etmek
        # icin cache'i bilerek atlayan bir surumle test edelim: get_or_create_animal'i
        # bos bir cache ile dogrudan cagiralim.
        logger = S.setup_logging(cfg.log_path)
        with Ctx(client, "sync-profile-id"):
            animal_id = S.get_or_create_animal(client, logger, "555", {}, "sync-profile-id")

        check("cakisan kupe icin var olan hayvan ID'si donduruldu", animal_id == existing_id, animal_id)
        check("ikinci bir hayvan olusturulmadi", len(client.db["animals"]) == 1, client.db["animals"])


def main():
    tests = [v for k, v in sorted(globals().items()) if k.startswith("test_")]
    print(f"Calistiriliyor: {len(tests)} test\n")
    for t in tests:
        print(f"{t.__name__}:")
        t()
        print()
    print(f"Sonuc: {PASSED} basarili, {FAILED} basarisiz")
    return 1 if FAILED else 0


if __name__ == "__main__":
    raise SystemExit(main())
